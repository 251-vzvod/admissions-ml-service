"""Build a manageable seed corpus from downloaded external datasets."""

from __future__ import annotations

import argparse
import csv
import json
from pathlib import Path
from typing import Any, Iterable
from zipfile import ZipFile


ROOT = Path(__file__).resolve().parents[1]
RAW_ROOT = ROOT / "data" / "external" / "raw"


def iter_persuade_rows(path: Path, limit: int) -> Iterable[dict[str, Any]]:
    with path.open("r", encoding="utf-8", newline="") as handle:
        reader = csv.DictReader(handle)
        for idx, row in enumerate(reader):
            if idx >= limit:
                break
            yield {
                "source_dataset": "persuade_2_train",
                "record_type": "student_argumentative_essay",
                "text": row.get("full_text", ""),
                "labels": {
                    "holistic_essay_score": row.get("holistic_essay_score"),
                    "competition_set": row.get("competition_set"),
                    "assignment": row.get("assignment"),
                },
                "metadata": {
                    "essay_id": row.get("essay_id"),
                    "essay_id_comp": row.get("essay_id_comp"),
                },
            }


def iter_asap2_rows(path: Path, limit: int) -> Iterable[dict[str, Any]]:
    with path.open("r", encoding="utf-8", newline="") as handle:
        reader = csv.DictReader(handle)
        for idx, row in enumerate(reader):
            if idx >= limit:
                break
            yield {
                "source_dataset": "learning_agency_asap_2",
                "record_type": "student_argumentative_essay",
                "text": row.get("full_text", ""),
                "labels": {
                    "holistic_essay_score": row.get("score"),
                    "assignment": row.get("assignment"),
                    "prompt_name": row.get("prompt_name"),
                },
                "metadata": {
                    "essay_id": row.get("essay_id"),
                },
            }


def iter_aide_rows(path: Path, limit: int) -> Iterable[dict[str, Any]]:
    with path.open("r", encoding="utf-8", newline="") as handle:
        reader = csv.DictReader(handle)
        for idx, row in enumerate(reader):
            if idx >= limit:
                break
            yield {
                "source_dataset": "learning_agency_aide",
                "record_type": "essay_authenticity_example",
                "text": row.get("text", ""),
                "labels": {
                    "generated": row.get("generated"),
                    "prompt_id": row.get("prompt_id"),
                },
                "metadata": {
                    "id": row.get("id"),
                },
            }


def iter_kazparc_rows(path: Path, limit: int) -> Iterable[dict[str, Any]]:
    with path.open("r", encoding="utf-8", newline="") as handle:
        reader = csv.DictReader(handle)
        text_columns = [
            column
            for column in (reader.fieldnames or [])
            if column.lower() not in {"id", "idx", "index", "domain", "source", "split"}
        ]
        for idx, row in enumerate(reader):
            if idx >= limit:
                break
            multilingual_payload = {
                column: value
                for column, value in row.items()
                if column in text_columns and isinstance(value, str) and value.strip()
            }
            yield {
                "source_dataset": "kazparc_all_entries",
                "record_type": "multilingual_parallel_text",
                "text": multilingual_payload.get("en")
                or multilingual_payload.get("english")
                or next(iter(multilingual_payload.values()), ""),
                "multilingual_texts": multilingual_payload,
                "metadata": {
                    "domain": row.get("domain"),
                },
            }


def iter_piilo_rows(path: Path, limit: int) -> Iterable[dict[str, Any]]:
    with path.open("r", encoding="utf-8") as handle:
        payload = json.load(handle)

    for idx, row in enumerate(payload):
        if idx >= limit:
            break
        full_text = row.get("full_text", "")
        if not isinstance(full_text, str) or not full_text.strip():
            continue
        yield {
            "source_dataset": "learning_agency_piilo",
            "record_type": "privacy_redaction_example",
            "text": full_text,
            "labels": {
                "document": row.get("document"),
                "pii_label_count": len(row.get("labels") or []),
            },
            "metadata": {
                "token_count": len(row.get("tokens") or []),
            },
        }


def iter_clear_rows(path: Path, limit: int) -> Iterable[dict[str, Any]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - dependency guard
        raise RuntimeError("openpyxl_required_for_clear_corpus") from exc

    workbook = load_workbook(filename=path, read_only=True, data_only=True)
    sheet = workbook.active
    rows = sheet.iter_rows(values_only=True)
    header = [str(cell).strip() if cell is not None else "" for cell in next(rows)]
    lowered = [item.lower() for item in header]

    excerpt_idx = lowered.index("excerpt") if "excerpt" in lowered else None
    id_idx = lowered.index("id") if "id" in lowered else None
    category_idx = lowered.index("category") if "category" in lowered else None
    year_idx = lowered.index("pub year") if "pub year" in lowered else None
    easiness_idx = lowered.index("bt easiness") if "bt easiness" in lowered else None

    count = 0
    for row in rows:
        if count >= limit:
            break
        excerpt = row[excerpt_idx] if excerpt_idx is not None and excerpt_idx < len(row) else None
        if not isinstance(excerpt, str) or not excerpt.strip():
            continue
        yield {
            "source_dataset": "clear_corpus",
            "record_type": "reading_passage",
            "text": excerpt,
            "labels": {
                "bt_easiness": row[easiness_idx] if easiness_idx is not None and easiness_idx < len(row) else None,
            },
            "metadata": {
                "id": row[id_idx] if id_idx is not None and id_idx < len(row) else None,
                "category": row[category_idx] if category_idx is not None and category_idx < len(row) else None,
                "pub_year": row[year_idx] if year_idx is not None and year_idx < len(row) else None,
            },
        }
        count += 1


def iter_mediasum_rows(path: Path, limit: int) -> Iterable[dict[str, Any]]:
    with ZipFile(path) as archive:
        member_name = next((name for name in archive.namelist() if name.endswith(".json")), None)
        if not member_name:
            return
        with archive.open(member_name, "r") as handle:
            payload = json.load(handle)
        for idx, row in enumerate(payload):
            if idx >= limit:
                break
            utterances = row.get("utt") or []
            text = "\n".join(item for item in utterances if isinstance(item, str))
            yield {
                "source_dataset": "mediasum",
                "record_type": "interview_transcript",
                "text": text,
                "labels": {
                    "summary": row.get("summary"),
                    "program": row.get("program"),
                },
                "metadata": {
                    "id": row.get("id"),
                    "title": row.get("title"),
                    "date": row.get("date"),
                    "url": row.get("url"),
                },
            }


def main() -> None:
    parser = argparse.ArgumentParser(description="Build seed corpus from external raw datasets")
    parser.add_argument("--output", default="data/external/seed_corpus.jsonl", help="Where to write the seed corpus")
    parser.add_argument("--persuade-limit", type=int, default=2500)
    parser.add_argument("--asap2-limit", type=int, default=2000)
    parser.add_argument("--aide-limit", type=int, default=2000)
    parser.add_argument("--kazparc-limit", type=int, default=5000)
    parser.add_argument("--piilo-limit", type=int, default=500)
    parser.add_argument("--clear-limit", type=int, default=2000)
    parser.add_argument("--mediasum-limit", type=int, default=0)
    args = parser.parse_args()

    output_path = ROOT / args.output
    output_path.parent.mkdir(parents=True, exist_ok=True)

    records: list[dict[str, Any]] = []

    persuade_path = RAW_ROOT / "persuade_2_train" / "persuade_2_train.csv"
    if persuade_path.exists():
        records.extend(iter_persuade_rows(persuade_path, limit=args.persuade_limit))

    asap2_path = RAW_ROOT / "learning_agency_asap_2" / "ASAP2_train_sourcetexts.csv"
    if asap2_path.exists():
        records.extend(iter_asap2_rows(asap2_path, limit=args.asap2_limit))

    aide_path = RAW_ROOT / "learning_agency_aide" / "train_essays.csv"
    if aide_path.exists():
        records.extend(iter_aide_rows(aide_path, limit=args.aide_limit))

    kazparc_path = RAW_ROOT / "kazparc_all_entries" / "kazparc_all_entries.csv"
    if kazparc_path.exists():
        records.extend(iter_kazparc_rows(kazparc_path, limit=args.kazparc_limit))

    piilo_path = RAW_ROOT / "learning_agency_piilo" / "train.json"
    if piilo_path.exists():
        records.extend(iter_piilo_rows(piilo_path, limit=args.piilo_limit))

    clear_path = RAW_ROOT / "clear_corpus" / "CLEAR_corpus_final.xlsx"
    if clear_path.exists():
        records.extend(iter_clear_rows(clear_path, limit=args.clear_limit))

    mediasum_path = RAW_ROOT / "mediasum" / "mediasum.zip"
    if mediasum_path.exists() and args.mediasum_limit > 0:
        records.extend(iter_mediasum_rows(mediasum_path, limit=args.mediasum_limit))

    with output_path.open("w", encoding="utf-8") as handle:
        for record in records:
            handle.write(json.dumps(record, ensure_ascii=False) + "\n")

    summary = {
        "output": str(output_path.relative_to(ROOT)),
        "record_count": len(records),
        "sources_present": sorted({record["source_dataset"] for record in records}),
    }
    summary_path = output_path.with_suffix(".summary.json")
    summary_path.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(summary, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
